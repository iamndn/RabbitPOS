#!/usr/bin/env python3
"""Analyze Excel files structure for migration planning."""
import openpyxl
import sys

def analyze_file(filename):
    print(f"\n{'='*60}")
    print(f"FILE: {filename}")
    print(f"{'='*60}")
    wb = openpyxl.load_workbook(filename)
    print(f"Sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n--- Sheet: '{sn}' ({ws.max_row} rows, {ws.max_column} cols) ---")
        headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        print(f"HEADERS: {headers}")
        print("\nSample rows:")
        for r in range(2, min(7, ws.max_row+1)):
            row = [ws.cell(r,c).value for c in range(1, ws.max_column+1)]
            print(f"  Row {r}: {row}")
        print(f"\nLast rows:")
        for r in range(max(2, ws.max_row-2), ws.max_row+1):
            row = [ws.cell(r,c).value for c in range(1, ws.max_column+1)]
            print(f"  Row {r}: {row}")

analyze_file('/opt/RabbitPOS/SoThuChi.xlsx')
analyze_file('/opt/RabbitPOS/DonHang.xlsx')
print("\nDone!")
