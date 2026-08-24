#!/usr/bin/env python3
"""
=============================================================================
RabbitPOS - Migrate từ SoBanHang App
=============================================================================
"""

import openpyxl
import psycopg2
import os
import sys
import re
from datetime import datetime, timezone, timedelta
from collections import defaultdict

# ============================================================================
# CONFIG
# ============================================================================

DB_CONFIG = {
    "host":     "localhost",
    "port":     5432,
    "user":     "admin",
    "password": "SGsurv9wi7uQRXXniYj0kg1V8TunrUnR",
    "dbname":   "rabbitpos",
}

SOTHU_CHI_FILE = "/opt/RabbitPOS/SoThuChi.xlsx"
DON_HANG_FILE  = "/opt/RabbitPOS/DonHang.xlsx"

VN_TZ = timezone(timedelta(hours=7))

PRODUCT_NAME_MAP = {
    "dưa hấu dứa": "DƯA HẤU + DỨA",
    "dua hau dua":  "DƯA HẤU + DỨA",
}

TX_CATEGORY_MAP = {
    "doanh thu bán hàng pos": "Doanh thu bán hàng POS",
    "mua nguyên vật liệu":    "Mua nguyên vật liệu",
    "thiết bị, dụng cụ, phần mềm": "Thiết bị, dụng cụ, phần mềm",
    "tiền đá viên":           "Tiền đá viên",
    "thuê nhà, mặt bằng, văn phòng": "Thuê nhà, mặt bằng, văn phòng",
    "chi phí marketing":      "Chi phí marketing",
    "điện, nước, internet":   "Điện, nước, internet",
    "thu nhập khác":          "Thu nhập khác",
}

# ============================================================================
# HELPERS
# ============================================================================

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

def log_warn(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] WARN: {msg}", flush=True)

def log_error(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] ERROR: {msg}", flush=True)

def log_ok(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] OK: {msg}", flush=True)

def make_aware(dt):
    if dt is None:
        return datetime.now(VN_TZ)
    if isinstance(dt, datetime):
        if dt.tzinfo is None:
            return dt.replace(tzinfo=VN_TZ)
        return dt.astimezone(VN_TZ)
    return datetime.now(VN_TZ)

def parse_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        clean = re.sub(r'[^\d.]', '', v)
        try:
            return float(clean) if clean else 0.0
        except:
            return 0.0
    return 0.0

def normalize_product_name(name):
    if not name:
        return ""
    name_lower = name.strip().lower()
    if name_lower in PRODUCT_NAME_MAP:
        return PRODUCT_NAME_MAP[name_lower]
    return name.strip().upper()

def generate_order_code(dt, order_id=0):
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=VN_TZ)
    ms = dt.microsecond // 1000
    # Use order_id as tiebreaker if needed
    suffix = (ms + order_id * 7) % 10000
    return f"ORD-{dt.strftime('%Y%m%d-%H%M%S')}-{suffix:04d}"

def normalize_fund(name, fund_map, fallback=1):
    if not name:
        return fallback
    nl = name.strip().lower()
    if nl in fund_map:
        return fund_map[nl]
    # Aliases
    aliases = {
        "tiền mặt": 1,
        "cash": 1,
        "tien mat": 1,
        "tiền mặt tại quầy": 1,
        "chuyển khoản": 2,
        "vietqr": 2,
        "bank": 2,
        "chuyển khoản vietqr": 2,
    }
    for k, v in aliases.items():
        if k in nl or nl in k:
            return v
    return fallback

# ============================================================================
# LOAD EXCEL
# ============================================================================

def load_sothu_chi(filepath):
    log(f"Đọc file SoThuChi: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    rows = []
    warnings = []
    
    for r in range(2, ws.max_row + 1):
        date_val   = ws.cell(r, 1).value
        type_val   = ws.cell(r, 2).value
        fund_val   = ws.cell(r, 3).value
        cat_val    = ws.cell(r, 4).value
        amount_val = ws.cell(r, 5).value
        note_val   = ws.cell(r, 6).value
        cashier_val= ws.cell(r, 7).value

        if not date_val and not amount_val:
            continue

        created_at = make_aware(date_val if isinstance(date_val, datetime) else datetime.now())
        
        type_str = str(type_val or "").lower()
        tx_type = "inflow" if ("thu" in type_str or "inflow" in type_str) else "outflow"
        
        fund_str = str(fund_val or "").strip()
        # Detect wrong column error
        if "thu tiền" in fund_str.lower() or "chi tiền" in fund_str.lower() or "inflow" in fund_str.lower():
            warnings.append(f"  Row {r}: Quỹ Tiền có giá trị lỗi '{fund_str}' → fallback Tiền mặt tại quầy")
            fund_str = "Tiền mặt tại quầy"
        
        amount = parse_num(amount_val)
        if amount <= 0:
            warnings.append(f"  Row {r}: Số tiền = 0 → bỏ qua")
            continue
        
        cat_str = str(cat_val or "Khác").strip()
        cat_str = TX_CATEGORY_MAP.get(cat_str.lower(), cat_str)
        
        rows.append({
            "row_idx": r,
            "created_at": created_at,
            "tx_type": tx_type,
            "fund_name": fund_str,
            "category": cat_str,
            "amount": amount,
            "description": str(note_val or "").strip(),
            "cashier": str(cashier_val or "").strip() or None,
        })
    
    log(f"  → {len(rows)} giao dịch, {len(warnings)} cảnh báo")
    for w in warnings:
        log_warn(w)
    return rows, warnings

def load_don_hang(filepath):
    log(f"Đọc file DonHang: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Lịch Sử Đơn Hàng']
    
    order_groups = defaultdict(list)
    order_seq = []
    warnings = []
    
    for r in range(2, ws.max_row + 1):
        old_code    = ws.cell(r, 1).value
        date_val    = ws.cell(r, 2).value
        cashier_val = ws.cell(r, 3).value
        payment_val = ws.cell(r, 4).value
        status_val  = ws.cell(r, 5).value
        product_val = ws.cell(r, 6).value
        size_val    = ws.cell(r, 7).value
        # col 8: Giá bán trước giảm (skip)
        # col 9: Chiết khấu sản phẩm (skip)
        price_val   = ws.cell(r, 10).value  # Giá bán sau cùng
        qty_val     = ws.cell(r, 11).value
        # col 12: Tổng tiền
        discount_val= ws.cell(r, 13).value
        shipping_val= ws.cell(r, 14).value
        surcharge_val=ws.cell(r, 15).value
        net_val     = ws.cell(r, 16).value  # Thực Thu
        note_val    = ws.cell(r, 17).value
        
        if not old_code and not product_val:
            continue
        
        old_code = str(old_code or "").strip()
        if not old_code:
            continue
        
        created_at = make_aware(date_val if isinstance(date_val, datetime) else datetime.now())
        
        status_str = str(status_val or "").strip().lower()
        status = "cancelled" if ("hủy" in status_str or "cancel" in status_str) else "completed"
        
        product_name = normalize_product_name(str(product_val or "").strip())
        variant_name = str(size_val or "Mặc định").strip() or "Mặc định"
        
        unit_price = parse_num(price_val)
        quantity = max(1, int(parse_num(qty_val) or 1))
        net_total = parse_num(net_val)
        total_amount = net_total if net_total > 0 else parse_num(ws.cell(r, 12).value)
        discount = parse_num(discount_val)
        shipping = parse_num(shipping_val)
        surcharge = parse_num(surcharge_val)
        fund_name = str(payment_val or "").strip()
        cashier = str(cashier_val or "").strip()
        note = str(note_val or "").strip()
        
        item = {
            "row_idx": r,
            "old_code": old_code,
            "created_at": created_at,
            "status": status,
            "product_name": product_name,
            "variant_name": variant_name,
            "unit_price": unit_price,
            "quantity": quantity,
            "total_amount": total_amount,
            "discount": discount,
            "shipping": shipping,
            "surcharge": surcharge,
            "fund_name": fund_name,
            "cashier": cashier,
            "note": note,
        }
        
        if old_code not in order_groups:
            order_seq.append(old_code)
        order_groups[old_code].append(item)
    
    log(f"  → {len(order_seq)} đơn hàng unique")
    return order_seq, order_groups, warnings

# ============================================================================
# CLEAR DATA
# ============================================================================

def clear_existing_data(conn):
    log("Xoá dữ liệu cũ...")
    with conn.cursor() as cur:
        cur.execute("DELETE FROM purchase_items")
        pi = cur.rowcount
        cur.execute("DELETE FROM order_items")
        oi = cur.rowcount
        cur.execute("DELETE FROM transactions")
        tx = cur.rowcount
        cur.execute("DELETE FROM orders")
        od = cur.rowcount
        cur.execute("UPDATE funds SET current_balance = 0")
        cur.execute("SELECT setval(pg_get_serial_sequence('orders','id'), 1, false)")
        cur.execute("SELECT setval(pg_get_serial_sequence('order_items','id'), 1, false)")
        cur.execute("SELECT setval(pg_get_serial_sequence('transactions','id'), 1, false)")
    conn.commit()
    log_ok(f"  Đã xoá: {od} orders, {oi} order_items, {tx} transactions, {pi} purchase_items")

# ============================================================================
# IMPORT TRANSACTIONS
# ============================================================================

def import_transactions(conn, rows, fund_map):
    log(f"Import {len(rows)} giao dịch từ SoThuChi...")
    imported = 0
    errors = []
    
    with conn.cursor() as cur:
        for row in rows:
            fund_id = normalize_fund(row['fund_name'], fund_map, 1)
            cat_name = row['category']
            tx_type = row['tx_type']
            cat_type = "inflow" if tx_type == "inflow" else "outflow"
            
            # Upsert transaction category
            cur.execute("""
                INSERT INTO transaction_categories (name, type, is_system)
                VALUES (%s, %s, false)
                ON CONFLICT DO NOTHING
            """, (cat_name, cat_type))
            
            try:
                cur.execute("""
                    INSERT INTO transactions 
                        (fund_id, transaction_type, category, amount, description,
                         cashier_name, created_by, created_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    fund_id,
                    tx_type,
                    cat_name,
                    row['amount'],
                    row['description'],
                    row['cashier'] or '',
                    row['cashier'] or 'system',
                    row['created_at'],
                ))
                imported += 1
            except Exception as e:
                errors.append(f"Row {row['row_idx']}: {e}")
                conn.rollback()
    
    conn.commit()
    log_ok(f"  Import OK: {imported} | Lỗi: {len(errors)}")
    for e in errors[:5]:
        log_error(f"  {e}")
    return imported, errors

# ============================================================================
# IMPORT ORDERS
# ============================================================================

def import_orders(conn, order_seq, order_groups, fund_map, variant_map):
    log(f"Import {len(order_seq)} đơn hàng từ DonHang...")
    
    imported_orders = 0
    imported_items = 0
    skipped_orders = 0
    new_products = []
    errors = []
    code_mapping = {}  # order_id -> {old_code, new_code, created_at}
    
    with conn.cursor() as cur:
        for old_code in order_seq:
            items_data = order_groups[old_code]
            if not items_data:
                continue
            
            first = items_data[0]
            created_at = first['created_at']
            status = first['status']
            fund_id = normalize_fund(first['fund_name'], fund_map, 1)
            cashier = first['cashier'] or ''
            note = first['note'] or None
            discount = first['discount']
            shipping = first['shipping']
            surcharge = first['surcharge']
            
            # Build order items
            subtotal = 0.0
            order_items_to_insert = []
            
            for item in items_data:
                prod_name = item['product_name']
                var_name = item['variant_name']
                lookup_key = prod_name.lower() + "_" + var_name.lower()
                
                variant_id = variant_map.get(lookup_key)
                
                if variant_id is None:
                    log_warn(f"  Không tìm thấy: '{prod_name}' / '{var_name}' → tạo mới")
                    
                    # Find or create product
                    cur.execute("SELECT id FROM products WHERE LOWER(name) = LOWER(%s) LIMIT 1", (prod_name,))
                    prod_row = cur.fetchone()
                    if prod_row:
                        prod_id = prod_row[0]
                    else:
                        cur.execute("""
                            INSERT INTO products (category_id, name, is_active, created_at, updated_at)
                            VALUES (1, %s, true, NOW(), NOW()) RETURNING id
                        """, (prod_name,))
                        prod_id = cur.fetchone()[0]
                        new_products.append(prod_name)
                    
                    cur.execute("""
                        INSERT INTO product_variants (product_id, variant_name, retail_price, cogs_price, is_active, created_at, updated_at)
                        VALUES (%s, %s, %s, 0, true, NOW(), NOW()) RETURNING id
                    """, (prod_id, var_name, item['unit_price'] or 0))
                    variant_id = cur.fetchone()[0]
                    variant_map[lookup_key] = variant_id
                    log_ok(f"    Đã tạo variant ID={variant_id} cho '{prod_name}' / '{var_name}'")
                
                unit_price = item['unit_price']
                qty = item['quantity']
                line_total = unit_price * qty
                subtotal += line_total
                
                order_items_to_insert.append((variant_id, qty, unit_price, line_total, created_at))
            
            # Calculate total
            total_amount = first['total_amount']
            if total_amount <= 0:
                total_amount = subtotal - discount + shipping + surcharge
            
            # Temporary code to avoid unique constraint during batch insert
            temp_code = f"TMP-{old_code}"
            
            try:
                cur.execute("""
                    INSERT INTO orders 
                        (order_code, status, subtotal, discount_amount, shipping_fee, surcharge,
                         total_amount, fund_id, cashier_name, created_by, note, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    temp_code, status, subtotal, discount, shipping, surcharge,
                    total_amount, fund_id, cashier, cashier or 'system',
                    note, created_at, created_at,
                ))
                order_id = cur.fetchone()[0]
                
                code_mapping[order_id] = {
                    "old_code": old_code,
                    "new_code": generate_order_code(created_at, order_id),
                    "created_at": created_at,
                }
                
                for vid, qty, up, lt, cat in order_items_to_insert:
                    cur.execute("""
                        INSERT INTO order_items 
                            (order_id, product_variant_id, quantity, unit_price, line_total,
                             selected_toppings, toppings_price, notes, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, %s, '[]', 0, '', %s, %s)
                    """, (order_id, vid, qty, up, lt, cat, cat))
                    imported_items += 1
                
                imported_orders += 1
                
            except Exception as e:
                conn.rollback()
                errors.append(f"Order {old_code}: {e}")
                skipped_orders += 1
                log_error(f"  Lỗi đơn {old_code}: {e}")
    
    conn.commit()
    log_ok(f"  Import OK: {imported_orders} đơn, {imported_items} items | Bỏ qua: {skipped_orders}")
    if new_products:
        log_warn(f"  Tạo mới {len(new_products)} sản phẩm: {new_products}")
    return code_mapping, errors

# ============================================================================
# RENAME ORDER CODES
# ============================================================================

def rename_order_codes(conn, code_mapping):
    log(f"Đổi {len(code_mapping)} mã đơn hàng...")
    renamed = 0
    
    with conn.cursor() as cur:
        # Check for duplicate new_codes and handle
        code_counts = {}
        for oid, info in code_mapping.items():
            nc = info['new_code']
            if nc in code_counts:
                # Resolve conflict: add order_id to suffix
                info['new_code'] = f"ORD-{info['created_at'].strftime('%Y%m%d-%H%M%S')}-{oid % 10000:04d}"
            code_counts[nc] = oid
        
        for order_id, info in code_mapping.items():
            cur.execute(
                "UPDATE orders SET order_code = %s WHERE id = %s",
                (info['new_code'], order_id)
            )
            renamed += 1
    
    conn.commit()
    log_ok(f"  Đổi tên: {renamed} mã đơn hàng")
    
    # Show sample
    with conn.cursor() as cur:
        cur.execute("SELECT order_code, created_at, status FROM orders ORDER BY created_at LIMIT 5")
        rows = cur.fetchall()
        log("  Mẫu mã đơn mới:")
        for r in rows:
            log(f"    {r[0]} | {r[1]} | {r[2]}")

# ============================================================================
# RECALCULATE BALANCES
# ============================================================================

def recalculate_fund_balances(conn):
    log("Tính lại fund balances...")
    with conn.cursor() as cur:
        cur.execute("""
            UPDATE funds f SET 
                current_balance = COALESCE((
                    SELECT SUM(CASE WHEN t.transaction_type = 'inflow' THEN t.amount ELSE -t.amount END)
                    FROM transactions t WHERE t.fund_id = f.id
                ), 0),
                updated_at = NOW()
        """)
        cur.execute("SELECT id, name, current_balance FROM funds ORDER BY id")
        funds = cur.fetchall()
    conn.commit()
    log_ok("Fund balances sau migration:")
    for f in funds:
        log(f"  Fund {f[0]} ({f[1]}): {f[2]:,.0f} VNĐ")

def reset_sequences(conn):
    tables = ["orders", "order_items", "transactions", "products", "product_variants"]
    with conn.cursor() as cur:
        for t in tables:
            cur.execute(f"""
                SELECT setval(pg_get_serial_sequence('{t}', 'id'),
                    COALESCE((SELECT MAX(id) FROM {t}), 1),
                    (SELECT MAX(id) IS NOT NULL FROM {t}))
            """)
    conn.commit()
    log_ok("Reset sequences OK")

# ============================================================================
# MAIN
# ============================================================================

def main():
    print("=" * 60)
    print("  RabbitPOS - Migrate Từ SoBanHang")
    print("=" * 60)
    
    # Connect DB
    log("Kết nối database...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        conn.autocommit = False
        log_ok("Kết nối thành công!")
    except Exception as e:
        log_error(f"Lỗi kết nối DB: {e}")
        sys.exit(1)
    
    try:
        # Load fund map from DB
        with conn.cursor() as cur:
            cur.execute("SELECT id, name, fund_type FROM funds")
            funds_data = cur.fetchall()
        
        fund_map = {}
        for fid, fname, ftype in funds_data:
            fund_map[fname.strip().lower()] = fid
            if ftype == 'cash':
                fund_map['tiền mặt'] = fid
                fund_map['cash'] = fid
                fund_map['tiền mặt tại quầy'] = fid
            elif ftype in ('bank', 'e-wallet'):
                fund_map['bank'] = fid
                fund_map['chuyển khoản'] = fid
                fund_map['chuyển khoản vietqr'] = fid
                fund_map['vietqr'] = fid
        
        log(f"Funds: {[(f[0], f[1]) for f in funds_data]}")
        
        # Load variant map from DB
        with conn.cursor() as cur:
            cur.execute("""
                SELECT pv.id, p.name, pv.variant_name
                FROM product_variants pv JOIN products p ON p.id = pv.product_id
            """)
            variants_data = cur.fetchall()
        
        variant_map = {}
        for vid, pname, vname in variants_data:
            key = pname.lower() + "_" + vname.lower()
            variant_map[key] = vid
        log(f"Variants trong DB: {len(variant_map)}")
        
        # === LOAD EXCEL ===
        print()
        tx_rows, tx_warns = load_sothu_chi(SOTHU_CHI_FILE)
        order_seq, order_groups, ord_warns = load_don_hang(DON_HANG_FILE)
        
        # === DRY RUN REPORT ===
        print()
        log("=== PREVIEW ===")
        cancelled_count = sum(1 for code in order_seq if order_groups[code][0]['status'] == 'cancelled')
        log(f"Sẽ import:")
        log(f"  - {len(tx_rows)} giao dịch thu chi từ SoThuChi")
        log(f"  - {len(order_seq)} đơn hàng ({cancelled_count} đã hủy) từ DonHang")
        log(f"  - Xoá dữ liệu hiện tại: orders, transactions, order_items, purchase_items")
        log(f"  - Giữ nguyên: products, categories")
        
        print()
        confirm = input(">>> Nhập 'YES' để bắt đầu migration: ").strip()
        if confirm != "YES":
            log("Đã huỷ!")
            return
        
        # === EXECUTE ===
        print()
        clear_existing_data(conn)
        
        print()
        tx_imported, tx_errors = import_transactions(conn, tx_rows, fund_map)
        
        print()
        code_mapping, ord_errors = import_orders(conn, order_seq, order_groups, fund_map, variant_map)
        
        print()
        rename_order_codes(conn, code_mapping)
        
        print()
        recalculate_fund_balances(conn)
        
        reset_sequences(conn)
        
        # === FINAL REPORT ===
        print()
        print("=" * 60)
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM orders")
            total_orders = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM order_items")
            total_items = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM transactions")
            total_tx = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM orders WHERE status = 'cancelled'")
            cancelled = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM products")
            total_prods = cur.fetchone()[0]
        
        print(f"""
MIGRATION HOÀN THÀNH!

  Orders nhập:         {total_orders} ({cancelled} đã hủy)
  Order items:         {total_items}
  Transactions nhập:   {total_tx}
  Sản phẩm trong DB:   {total_prods}

  Lỗi SoThuChi:        {len(tx_errors)}
  Lỗi DonHang:         {len(ord_errors)}
""")
        
        if tx_errors + ord_errors:
            log_warn("Chi tiết lỗi:")
            for e in (tx_errors + ord_errors)[:20]:
                log_warn(f"  {e}")
        
        log_ok("Done!")
        
    except Exception as e:
        log_error(f"Lỗi: {e}")
        import traceback
        traceback.print_exc()
        conn.rollback()
        sys.exit(1)
    finally:
        conn.close()

if __name__ == "__main__":
    main()
