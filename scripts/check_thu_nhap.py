#!/usr/bin/env python3
import openpyxl, sys
wb = openpyxl.load_workbook('/opt/RabbitPOS/SoThuChi.xlsx')
ws = wb.active
for r in range(2, ws.max_row+1):
    cat = ws.cell(r,4).value
    if cat and 'thu nhập' in str(cat).lower():
        print(f"Row {r}: Type={ws.cell(r,2).value!r}, Fund={ws.cell(r,3).value!r}, Cat={cat!r}, Amt={ws.cell(r,5).value}, Note={ws.cell(r,6).value!r}", flush=True)
print("done", flush=True)
