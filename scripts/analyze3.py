#!/usr/bin/env python3
import openpyxl
import sys

wb = openpyxl.load_workbook('/opt/RabbitPOS/DonHang.xlsx')
ws = wb['Lịch Sử Đơn Hàng']
print("Col mapping:", flush=True)
for c in range(1, ws.max_column+1):
    print(f"  Col {c}: {ws.cell(1,c).value!r}", flush=True)

print("Row 2:", flush=True)
for c in range(1, ws.max_column+1):
    print(f"  Col {c}: {ws.cell(2,c).value!r}", flush=True)

# Check Dua hau
print("Dua hau variants:", flush=True)
seen = set()
for r in range(2, ws.max_row+1):
    v = ws.cell(r,6).value
    if v and v not in seen:
        seen.add(v)
        
# List unique product+size combos
combo = set()
for r in range(2, ws.max_row+1):
    prod = ws.cell(r,6).value
    size = ws.cell(r,7).value
    if prod:
        combo.add((prod, size))
print("Unique prod+size combos:", flush=True)
for item in sorted(combo):
    print(f"  {item}", flush=True)

print("Done", flush=True)
