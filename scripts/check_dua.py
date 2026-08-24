#!/usr/bin/env python3
import openpyxl, sys
wb = openpyxl.load_workbook('/opt/RabbitPOS/DonHang.xlsx')
ws = wb['Lịch Sử Đơn Hàng']
for r in range(2, ws.max_row+1):
    p = ws.cell(r,6).value
    if p and 'ưa hấu' in str(p).lower() and 'ứa' in str(p).lower() and '+' not in str(p):
        print(f"Row {r}: Prod={p!r}, Size={ws.cell(r,7).value!r}, Price={ws.cell(r,10).value}, Order={ws.cell(r,1).value}", flush=True)
print("done", flush=True)
