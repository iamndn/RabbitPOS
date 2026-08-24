#!/usr/bin/env python3
import openpyxl
from collections import Counter
import sys

print("START", flush=True)

# SoThuChi
wb = openpyxl.load_workbook('/opt/RabbitPOS/SoThuChi.xlsx')
ws = wb['SoThuChi']
print(f"SoThuChi rows: {ws.max_row - 1}", flush=True)

types = Counter()
funds = Counter()
categories = Counter()

for r in range(2, ws.max_row+1):
    types[ws.cell(r,2).value] += 1
    funds[ws.cell(r,3).value] += 1
    categories[ws.cell(r,4).value] += 1

print(f"Types: {dict(types)}", flush=True)
print(f"Funds: {dict(funds)}", flush=True)
print(f"Categories: {dict(categories)}", flush=True)

wb2 = openpyxl.load_workbook('/opt/RabbitPOS/DonHang.xlsx')
ws2 = wb2['Lịch Sử Đơn Hàng']
print(f"\nDonHang rows: {ws2.max_row - 1}", flush=True)

statuses = Counter()
payments = Counter()
unique_orders = set()

for r in range(2, ws2.max_row+1):
    unique_orders.add(ws2.cell(r,1).value)
    statuses[ws2.cell(r,5).value] += 1
    payments[ws2.cell(r,4).value] += 1

products = sorted(set(ws2.cell(r,6).value for r in range(2, ws2.max_row+1)))
print(f"Unique orders: {len(unique_orders)}", flush=True)
print(f"Statuses: {dict(statuses)}", flush=True)
print(f"Payments: {dict(payments)}", flush=True)
print(f"Unique products ({len(products)}): {products}", flush=True)
print("END", flush=True)
